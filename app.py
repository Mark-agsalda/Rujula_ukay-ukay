import os
import tempfile
from datetime import datetime
from flask import Flask, render_template_string, request, redirect, url_for, send_file
import mysql.connector

# Excel export support
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'ukay_live_secret_key_2026')

# MySQL Database Configuration
MYSQL_HOST = os.environ.get('MYSQL_HOST', 'localhost')
MYSQL_USER = os.environ.get('MYSQL_USER', 'root')
MYSQL_PASSWORD = os.environ.get('MYSQL_PASSWORD', '')
MYSQL_DB = os.environ.get('MYSQL_DB', 'ukay_inventory')
MYSQL_PORT = int(os.environ.get('MYSQL_PORT', 3306))

def get_db_connection():
    """Connects to MySQL server and database with Philippine Timezone (+08:00)."""
    conn = mysql.connector.connect(
        host=MYSQL_HOST,
        user=MYSQL_USER,
        password=MYSQL_PASSWORD,
        database=MYSQL_DB,
        port=MYSQL_PORT
    )
    cursor = conn.cursor()
    cursor.execute("SET time_zone = '+08:00';")
    cursor.close()
    return conn

def init_db():
    """Creates database and table if they do not exist."""
    try:
        conn = mysql.connector.connect(
            host=MYSQL_HOST,
            user=MYSQL_USER,
            password=MYSQL_PASSWORD,
            port=MYSQL_PORT
        )
        cursor = conn.cursor()
        cursor.execute(f"CREATE DATABASE IF NOT EXISTS `{MYSQL_DB}`")
        conn.commit()
        cursor.close()
        conn.close()

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ukay_inventory (
                id INT AUTO_INCREMENT PRIMARY KEY,
                mine_code VARCHAR(50) NOT NULL,
                item_description VARCHAR(255) NOT NULL,
                price DECIMAL(10,2) NOT NULL,
                buyer_name VARCHAR(100) NOT NULL,
                status VARCHAR(50) DEFAULT 'Mined',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.commit()
        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        print(f"MySQL Connection/Init Error: {err}")

with app.app_context():
    init_db()

HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Ukay-Ukay Live Inventory</title>
    <style>
        :root { --bg: #f8fafc; --card: #ffffff; --primary: #2563eb; --accent: #16a34a; --border: #e2e8f0; }
        body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; background: var(--bg); margin: 0; padding: 20px; color: #1e293b; }
        .container { max-width: 1100px; margin: 0 auto; }
        .header-bar { display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px; flex-wrap: wrap; gap: 10px; }
        h1 { margin: 0; color: #0f172a; font-size: 1.6rem; }
        .grid { display: grid; grid-template-columns: 320px 1fr; gap: 20px; }
        @media (max-width: 800px) { .grid { grid-template-columns: 1fr; } }
        .card { background: var(--card); padding: 20px; border-radius: 10px; border: 1px solid var(--border); box-shadow: 0 1px 3px rgba(0,0,0,0.05); margin-bottom: 15px; }
        .form-group { margin-bottom: 12px; }
        label { display: block; font-size: 0.85rem; font-weight: 600; margin-bottom: 4px; }
        input, select { width: 100%; padding: 8px 10px; border: 1px solid var(--border); border-radius: 6px; box-sizing: border-box; font-size: 0.95rem; }
        button, .btn { background: var(--primary); color: white; border: none; padding: 10px; width: 100%; border-radius: 6px; font-weight: 600; cursor: pointer; font-size: 0.95rem; text-decoration: none; display: inline-block; text-align: center; box-sizing: border-box; }
        button:hover, .btn:hover { opacity: 0.9; }
        .btn-excel { background: #16a34a; width: auto; padding: 8px 16px; font-size: 0.85rem; }
        .btn-filter { background: #475569; padding: 8px 12px; font-size: 0.85rem; }
        .btn-reset { background: #94a3b8; padding: 8px 12px; font-size: 0.85rem; color: white; text-decoration: none; border-radius: 6px; }
        .filter-form { display: flex; gap: 10px; align-items: flex-end; flex-wrap: wrap; }
        .filter-form .form-group { margin-bottom: 0; flex: 1; min-width: 130px; }
        .stats { display: flex; gap: 15px; margin-bottom: 15px; }
        .stat-box { background: #eff6ff; border: 1px solid #bfdbfe; padding: 12px; border-radius: 8px; flex: 1; text-align: center; }
        .stat-box .num { font-size: 1.4rem; font-weight: 700; color: #1e40af; }
        table { width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 0.9rem; }
        th, td { text-align: left; padding: 10px; border-bottom: 1px solid var(--border); }
        th { background: #f1f5f9; font-weight: 600; }
        .badge { padding: 3px 8px; border-radius: 12px; font-size: 0.75rem; font-weight: 600; }
        .badge-Mined { background: #fef3c7; color: #92400e; }
        .badge-Paid { background: #dcfce7; color: #166534; }
        .badge-Shipped { background: #e0e7ff; color: #3730a3; }
        .badge-Cancelled { background: #fee2e2; color: #991b1b; }
        .action-btns { display: flex; gap: 6px; }
        .btn-sm { padding: 4px 8px; font-size: 0.75rem; width: auto; border-radius: 4px; text-decoration: none; }
        .btn-edit { background: #0284c7; color: white; }
        .btn-del { background: #ef4444; color: white; }
        .date-col { font-size: 0.8rem; color: #64748b; }
    </style>
</head>
<body>

<div class="container">
    <div class="header-bar">
        <h1>👕 Ukay-Ukay Live Inventory Tracker</h1>
        {% if has_excel %}
            <a href="/export/excel?search_buyer={{ search_buyer }}&filter_date={{ filter_date }}&filter_month={{ filter_month }}&filter_year={{ filter_year }}" class="btn btn-excel">📊 Export View to Excel</a>
        {% endif %}
    </div>

    <!-- Filter & Search Bar -->
    <div class="card">
        <h3>🔍 Search & Filter Records</h3>
        <form action="/" method="GET" class="filter-form">
            <div class="form-group" style="flex: 1.5;">
                <label>Search Buyer Name / Handle</label>
                <input type="text" name="search_buyer" placeholder="e.g. Maria or @MariaClara" value="{{ search_buyer }}">
            </div>
            <div class="form-group">
                <label>Specific Date</label>
                <input type="date" name="filter_date" value="{{ filter_date }}">
            </div>
            <div class="form-group">
                <label>Month</label>
                <select name="filter_month">
                    <option value="">-- Any Month --</option>
                    <option value="1" {% if filter_month == '1' %}selected{% endif %}>January</option>
                    <option value="2" {% if filter_month == '2' %}selected{% endif %}>February</option>
                    <option value="3" {% if filter_month == '3' %}selected{% endif %}>March</option>
                    <option value="4" {% if filter_month == '4' %}selected{% endif %}>April</option>
                    <option value="5" {% if filter_month == '5' %}selected{% endif %}>May</option>
                    <option value="6" {% if filter_month == '6' %}selected{% endif %}>June</option>
                    <option value="7" {% if filter_month == '7' %}selected{% endif %}>July</option>
                    <option value="8" {% if filter_month == '8' %}selected{% endif %}>August</option>
                    <option value="9" {% if filter_month == '9' %}selected{% endif %}>September</option>
                    <option value="10" {% if filter_month == '10' %}selected{% endif %}>October</option>
                    <option value="11" {% if filter_month == '11' %}selected{% endif %}>November</option>
                    <option value="12" {% if filter_month == '12' %}selected{% endif %}>December</option>
                </select>
            </div>
            <div class="form-group">
                <label>Year</label>
                <input type="number" name="filter_year" placeholder="e.g. 2026" value="{{ filter_year }}">
            </div>
            <div style="display: flex; gap: 6px;">
                <button type="submit" class="btn btn-filter">Search</button>
                <a href="/" class="btn-reset">Reset</a>
            </div>
        </form>
    </div>

    <div class="grid">
        <!-- Input / Edit Form -->
        <div class="card">
            <h3>{% if edit_item %}✏️ Edit Record{% else %}⚡ Quick Add Mined Item{% endif %}</h3>
            <form action="{% if edit_item %}/update/{{ edit_item['id'] }}{% else %}/add{% endif %}" method="POST">
                <div class="form-group">
                    <label>Mine Code / Tag #</label>
                    <input type="text" name="mine_code" placeholder="e.g. M01" value="{{ edit_item['mine_code'] if edit_item else '' }}" required autofocus>
                </div>
                <div class="form-group">
                    <label>Item Description</label>
                    <input type="text" name="item_description" placeholder="e.g. Denim Jacket" value="{{ edit_item['item_description'] if edit_item else '' }}" required>
                </div>
                <div class="form-group">
                    <label>Price (₱)</label>
                    <input type="number" step="0.01" name="price" placeholder="150" value="{{ edit_item['price'] if edit_item else '' }}" required>
                </div>
                <div class="form-group">
                    <label>Buyer Name / Handle</label>
                    <input type="text" name="buyer_name" placeholder="e.g. @MariaClara" value="{{ edit_item['buyer_name'] if edit_item else '' }}" required>
                </div>
                <div class="form-group">
                    <label>Status</label>
                    <select name="status">
                        <option value="Mined" {% if edit_item and edit_item['status']=='Mined' %}selected{% endif %}>Mined (Unpaid)</option>
                        <option value="Paid" {% if edit_item and edit_item['status']=='Paid' %}selected{% endif %}>Paid</option>
                        <option value="Shipped" {% if edit_item and edit_item['status']=='Shipped' %}selected{% endif %}>Shipped</option>
                        <option value="Cancelled" {% if edit_item and edit_item['status']=='Cancelled' %}selected{% endif %}>Cancelled</option>
                    </select>
                </div>
                <button type="submit">{% if edit_item %}Update Record{% else %}Save Mined Item{% endif %}</button>
                {% if edit_item %}
                    <a href="/" style="display:block; text-align:center; margin-top:8px; font-size:0.85rem; color:#64748b; text-decoration:none;">Cancel Edit</a>
                {% endif %}
            </form>
        </div>

        <!-- Inventory Summary & List -->
        <div>
            <div class="stats">
                <div class="stat-box">
                    <div>Total Items</div>
                    <div class="num">{{ items|length }}</div>
                </div>
                <div class="stat-box">
                    <div>Total Revenue</div>
                    <div class="num">₱{{ "%.2f"|format(total_val) }}</div>
                </div>
            </div>

            <div class="card">
                <h3>📋 Recorded Items</h3>
                <table>
                    <thead>
                        <tr>
                            <th>Date Mined</th>
                            <th>Code</th>
                            <th>Description</th>
                            <th>Price</th>
                            <th>Buyer</th>
                            <th>Status</th>
                            <th>Actions</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for item in items %}
                        <tr>
                            <td class="date-col">{{ item['created_at'].strftime('%Y-%m-%d %H:%M') if item['created_at'] else 'N/A' }}</td>
                            <td><strong>{{ item['mine_code'] }}</strong></td>
                            <td>{{ item['item_description'] }}</td>
                            <td>₱{{ "%.2f"|format(item['price']) }}</td>
                            <td>{{ item['buyer_name'] }}</td>
                            <td><span class="badge badge-{{ item['status'] }}">{{ item['status'] }}</span></td>
                            <td class="action-btns">
                                <a href="/edit/{{ item['id'] }}?search_buyer={{ search_buyer }}&filter_date={{ filter_date }}&filter_month={{ filter_month }}&filter_year={{ filter_year }}" class="btn-sm btn-edit">Edit</a>
                                <a href="/delete/{{ item['id'] }}" class="btn-sm btn-del" onclick="return confirm('Delete record?')">X</a>
                            </td>
                        </tr>
                        {% else %}
                        <tr>
                            <td colspan="7" style="text-align: center; color: #94a3b8; padding: 20px;">No items found for this selection.</td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
        </div>
    </div>
</div>

</body>
</html>
'''

def fetch_filtered_items(search_buyer, f_date, f_month, f_year):
    """Utility function to build dynamic SQL queries based on search and date filters."""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    query = "SELECT * FROM ukay_inventory WHERE 1=1"
    params = []

    if search_buyer:
        query += " AND buyer_name LIKE %s"
        params.append(f"%{search_buyer.strip()}%")
    if f_date:
        query += " AND DATE(created_at) = %s"
        params.append(f_date)
    if f_month:
        query += " AND MONTH(created_at) = %s"
        params.append(f_month)
    if f_year:
        query += " AND YEAR(created_at) = %s"
        params.append(f_year)

    query += " ORDER BY id DESC"
    cursor.execute(query, tuple(params))
    items = cursor.fetchall()
    cursor.close()
    conn.close()
    return items

@app.route('/')
def index():
    search_buyer = request.args.get('search_buyer', '')
    filter_date = request.args.get('filter_date', '')
    filter_month = request.args.get('filter_month', '')
    filter_year = request.args.get('filter_year', '')

    items = fetch_filtered_items(search_buyer, filter_date, filter_month, filter_year)
    total_val = sum(float(item['price']) for item in items if item['status'] != 'Cancelled')

    return render_template_string(
        HTML_TEMPLATE,
        items=items,
        total_val=total_val,
        edit_item=None,
        has_excel=HAS_OPENPYXL,
        search_buyer=search_buyer,
        filter_date=filter_date,
        filter_month=filter_month,
        filter_year=filter_year
    )

@app.route('/add', methods=['POST'])
def add_item():
    code = request.form['mine_code']
    desc = request.form['item_description']
    price = float(request.form['price'])
    buyer = request.form['buyer_name']
    status = request.form['status']

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO ukay_inventory (mine_code, item_description, price, buyer_name, status)
        VALUES (%s, %s, %s, %s, %s)
    ''', (code, desc, price, buyer, status))
    conn.commit()
    cursor.close()
    conn.close()
    return redirect(url_for('index'))

@app.route('/edit/<int:item_id>')
def edit_item(item_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute('SELECT * FROM ukay_inventory WHERE id = %s', (item_id,))
    edit_item = cursor.fetchone()
    cursor.close()
    conn.close()

    search_buyer = request.args.get('search_buyer', '')
    filter_date = request.args.get('filter_date', '')
    filter_month = request.args.get('filter_month', '')
    filter_year = request.args.get('filter_year', '')

    items = fetch_filtered_items(search_buyer, filter_date, filter_month, filter_year)
    total_val = sum(float(item['price']) for item in items if item['status'] != 'Cancelled')

    return render_template_string(
        HTML_TEMPLATE,
        items=items,
        total_val=total_val,
        edit_item=edit_item,
        has_excel=HAS_OPENPYXL,
        search_buyer=search_buyer,
        filter_date=filter_date,
        filter_month=filter_month,
        filter_year=filter_year
    )

@app.route('/update/<int:item_id>', methods=['POST'])
def update_item(item_id):
    code = request.form['mine_code']
    desc = request.form['item_description']
    price = float(request.form['price'])
    buyer = request.form['buyer_name']
    status = request.form['status']

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE ukay_inventory
        SET mine_code=%s, item_description=%s, price=%s, buyer_name=%s, status=%s
        WHERE id=%s
    ''', (code, desc, price, buyer, status, item_id))
    conn.commit()
    cursor.close()
    conn.close()
    return redirect(url_for('index'))

@app.route('/delete/<int:item_id>')
def delete_item(item_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM ukay_inventory WHERE id = %s', (item_id,))
    conn.commit()
    cursor.close()
    conn.close()
    return redirect(url_for('index'))

@app.route('/export/excel')
def export_excel():
    if not HAS_OPENPYXL:
        return "openpyxl module is missing in requirements.txt", 500

    search_buyer = request.args.get('search_buyer', '')
    filter_date = request.args.get('filter_date', '')
    filter_month = request.args.get('filter_month', '')
    filter_year = request.args.get('filter_year', '')

    items = fetch_filtered_items(search_buyer, filter_date, filter_month, filter_year)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ukay Inventory"

    headers = ["ID", "Date Mined", "Mine Code", "Description", "Price (PHP)", "Buyer Name", "Status"]
    ws.append(headers)

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for item in items:
        ws.append([
            item['id'],
            str(item['created_at']),
            item['mine_code'],
            item['item_description'],
            float(item['price']),
            item['buyer_name'],
            item['status']
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, "ukay_inventory_export.xlsx")
    wb.save(file_path)

    return send_file(file_path, as_attachment=True, download_name=f"ukay_inventory_{datetime.now().strftime('%Y%m%d')}.xlsx")

if __name__ == '__main__':
    app.run(debug=True)
